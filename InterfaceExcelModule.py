#!/usr/bin/env python 
# -*- coding:utf-8 -*-

# __author__ = 'Hugo Chen'
# __time__   = '2019/1/14'
# God helps who help themselves!

import openpyxl   # excel操作模块
import openpyxl.styles    # excel样式调整

'''
EXCEL file 操作: 新建,打开,保存,查询
'''
# 新建一个PPT:
def createNewExcelFile():
    newExcel = openpyxl.Workbook()
    return newExcel

# 打开EXCEL:
def openExcelFile(path):
    openExcel = openpyxl.load_workbook(path)
    return openExcel

# 保存EXCEL:
def saveAs(openExcel, path):
    openExcel.save(path)

# 获取所有的sheet的数量:
def getSheetNum(excelBook):
    return len(excelBook.sheetnames)

# 获取所有的sheet的名称:
def getSheetNames(excelBook):
    return excelBook.sheetnames

'''
sheet 操作: 新建,打开,删除
'''
# 新建sheet表:
def createNewSheet(excelBook,title=None,index=None):
    sheet = excelBook.create_sheet(title,index)
    return sheet
# 删除sheet表:
def removeSheet(excelBook,index):
    sheet = excelBook.remove(excelBook[getSheetNames[index]])
    return sheet
# 获取sheet表:
def getOneSheet(excelBook,index):
    sheet = excelBook[getSheetNames[index]]
    return sheet
# 获取活动的sheet表:
def getActiveSheet(excelBook):
    return excelBook.active
# 通过名称列index转成名称
# def getColNameByIndex(index):

'''
读取数据
'''
# 获取单元格行数:
def getRowCount(sheet):
    return sheet.max_row
# 获取单元格列数:
def getCloumnCount(sheet):
    return sheet.max_column
# 获取单元格行名:
def getCellRowName(sheet,cellName):
    return sheet[cellName].row
# 获取单元格列名:
def getCellColumnName(sheet,cellName):
    return sheet[cellName].column
# 获取单元格数据:name:
def getCellValueByName(sheet,cellName):
    return sheet[cellName].value
# 获取单元格数据:row,col:
def getCellValue(sheet,row,column):
    return sheet.cell(row,column).value
# 获取某行所有数据
def getRowValueList(sheet,rowIdx):
    list_row = []
    for i in sheet[rowIdx]:
        list_row.append(i.value)
    return list_row
# 获取某列所有数据
def getColumnValueList(sheet,colName):
    list_row = []
    for i in sheet[colName]:
        list_row.append(i.value)
    return list_row
# 读取所有数据
def getAllValue(sheet):
    list_all= []
    for i in sheet.rows:    # i是一行，cell的list
        for j in i:    # j是cell
            list_all.append(j.value)
    return list_all
'''
写入数据
'''
# 向cell中输出
def setCellValueByName(sheet,cellName,valueTxt):
    sheet[cellName] = valueTxt
# 向cell中输出
def setCellValue(sheet,row,col,valueTxt):
    sheet.cell(row=row,column=col,value=valueTxt)
'''
表格样式调整
'''
#设置行高
def setRowHeight(sheet,rowIndex,height):
    sheet.row_dimensions[rowIndex].height = height
# 设置列宽
def setColumnWidth(sheet, colName, width):
    sheet.column_dimensions[colName].width = width
# 设置某个单元格的排列方式
def setCellAlign(sheet,cellName,horTxt='center',VerTxt='center'):
    sheet[cellName].alignment = openpyxl.styles.Alignment(horizontal=horTxt,vertical=VerTxt)
# 设置某个单元格的样式
def setCellFont(sheet, cellName, name="宋体", size=24, isBold=True, col="00CCFF", isItalic=False):
    sheet[cellName].font = openpyxl.styles.Font(name=name, size=size, bold=isBold, color=col, italic=isItalic)

if __name__ == '__main__':
    print("This is main function!")

    print("Application finished!")
